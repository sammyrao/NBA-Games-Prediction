# By Sammy Rao
import joblib
from openpyxl.xml.constants import MAX_ROW
import pandas as pd
import data_retriever as dr
from datetime import datetime, timedelta
from prettytable import PrettyTable
import openpyxl

# Load the model PKL file
model = joblib.load('data/models/model.pkl')

# Load the scaler PKL file
scaler = joblib.load('data/models/scaler.pkl')

# Function to start the prediction by getting the games schedule
def get_games_schedule(game_date, season):

    # Get the games schedule
    games_schedule_df = dr.retrieve_games_schedule(game_date)

    # Read the seasons CSV file
    seasons_CSV_df = pd.read_csv('data/seasons/seasons.csv')

    # Get the starting date of the season
    season_start_date = seasons_CSV_df.loc[seasons_CSV_df['SEASON'] == season, 'START_DATE'].values[0]

    # Format 'game_date' as a date and get the date before it
    game_date_formatted = datetime.strptime(game_date, '%m/%d/%Y').date()
    date_before_game = game_date_formatted - timedelta(days = 1)

    # Get the league standings of the teams
    league_standings_df = dr.retrieve_league_standings()

    # Prepare the data and proceed to prediction
    prepare_prediction_data(games_schedule_df, game_date, season_start_date, date_before_game, season, league_standings_df)

# Function to prepare the data for prediction
def prepare_prediction_data(games_schedule_df, game_date, season_start_date, date_before_game, season, league_standings_df):

    # Table to hold and display the predictions
    predictions = PrettyTable()

    # Set the title for the table
    predictions.title = 'Predictions for ' + game_date + ' fixtures'

    # Column names for the table
    predictions.field_names = ['#', 'Home Team | Visitor Team', 'League Record', 
                               'Home Record', 'Road Record', 'Current Streak', 'Playoff Rank', 
                               'Win Percentage', 'Predicted Winner', 'Probability']

    # Align the columns of the table
    predictions.align['#'] = 'c'
    predictions.align['Home Team | Visitor Team'] = 'l'
    predictions.align['League Record'] = 'c'
    predictions.align['Home Record'] = 'c'
    predictions.align['Road Record'] = 'c'
    predictions.align['Current Streak'] = 'c'   
    predictions.align['Playoff Rank'] = 'c'
    predictions.align['Win Percentage'] = 'c'
    predictions.align['Predicted Winner'] = 'c'
    predictions.align['Probability'] = 'c'

    # Add a horizontal line for every row
    predictions.hrules = 1

    # List to hold the prediction rows to be added to the testing Excel file
    # for testing the accuracy of the predictions or for further analysis
    predictions_list = []

    # Loop through the games schedule and get the stats for the teams
    for index, row in games_schedule_df.iterrows():

        # Get the IDs and nicknames of the teams
        team_a_id = row['HOME_TEAM_ID']
        team_a_nickname = row['HOME_TEAM_NICKNAME']
        team_b_id = row['VISITOR_TEAM_ID']
        team_b_nickname = row['VISITOR_TEAM_NICKNAME']

        print('Retrieving stats for Game ' + str(index + 1) + ' and predicting outcome . . .')

        team_a_stats = []
        team_b_stats = []

        # Get the stats for the teams
        team_a = dr.retrieve_team_stats(team_a_id, team_a_nickname, season_start_date, date_before_game, season)
        team_b = dr.retrieve_team_stats(team_b_id, team_b_nickname, season_start_date, date_before_game, season)
        
        # Append the dictionaries to the lists
        team_a_stats.append(team_a)
        team_b_stats.append(team_b)

        # Convert the lists into data frames
        team_a_stats_df = pd.DataFrame(team_a_stats)
        team_b_stats_df = pd.DataFrame(team_b_stats)

        # Get the stats differential onto a new data frame
        combined_stats_df = pd.DataFrame(team_a_stats_df.values - team_b_stats_df.values)
        # Give names to the columns
        combined_stats_df.columns = team_a_stats_df.columns
        # Add a prefix to the columns
        combined_stats_df = combined_stats_df.add_suffix('_DIFF')

        # Since 'team_a' is the home team, add a column with '1'
        combined_stats_df['HOME_TEAM'] = 1

        # Add the 'WINNER' column temporarily to avoid a value error when normalizing
        combined_stats_df['WINNER'] = 0

        # Fit and scale the data for normalization
        normalized_stats = scaler.transform(combined_stats_df)
        # Put the normalized data into a data frame
        normalized_stats_df = pd.DataFrame(normalized_stats)
        # Give names to the columns
        normalized_stats_df.columns = combined_stats_df.columns

        # Obtain the prediction result and add the row to the predictions table
        # Pass only the first 19 columns for the prediction (Leave out the 'WINNER' column)
        prediction = predict_game_outcome(team_a_nickname, team_b_nickname, normalized_stats_df.iloc[:, 0:19])
        
        # Extract the predicted winner from the returned result
        predicted_winner = prediction['Winner'][0]
        # Extract the probability from the returned result
        probability = prediction['Probability'][0]

        # Get the league records of the teams
        team_a_league_record = league_standings_df.loc[league_standings_df['TEAM_NAME'] == team_a_nickname, 'LEAGUE_RECORD'].values[0]
        team_b_league_record  = league_standings_df.loc[league_standings_df['TEAM_NAME'] == team_b_nickname, 'LEAGUE_RECORD'].values[0]
        # Remove white spaces
        team_a_league_record = team_a_league_record.strip()
        team_b_league_record = team_b_league_record.strip()

        # Get the home records of the teams
        team_a_home_record = league_standings_df.loc[league_standings_df['TEAM_NAME'] == team_a_nickname, 'HOME_RECORD'].values[0]
        team_b_home_record = league_standings_df.loc[league_standings_df['TEAM_NAME'] == team_b_nickname, 'HOME_RECORD'].values[0]
        # Remove white spaces
        team_a_home_record = team_a_home_record.strip()
        team_b_home_record = team_b_home_record.strip()

        # Get the road records of the teams
        team_a_road_record = league_standings_df.loc[league_standings_df['TEAM_NAME'] == team_a_nickname, 'ROAD_RECORD'].values[0]
        team_b_road_record = league_standings_df.loc[league_standings_df['TEAM_NAME'] == team_b_nickname, 'ROAD_RECORD'].values[0]
        # Remove white spaces
        team_a_road_record = team_a_road_record.strip()
        team_b_road_record = team_b_road_record.strip()

        # Get the current streaks of the teams
        team_a_current_streak = league_standings_df.loc[league_standings_df['TEAM_NAME'] == team_a_nickname, 'CURRENT_STREAK'].values[0]
        team_b_current_streak = league_standings_df.loc[league_standings_df['TEAM_NAME'] == team_b_nickname, 'CURRENT_STREAK'].values[0]             
        # Remove white spaces
        team_a_current_streak = team_a_current_streak.strip()
        team_b_current_streak = team_b_current_streak.strip()

        # Get the playoff ranks of the teams
        team_a_playoff_rank = league_standings_df.loc[league_standings_df['TEAM_NAME'] == team_a_nickname, 'PLAYOFF_RANK'].values[0]
        team_b_playoff_rank = league_standings_df.loc[league_standings_df['TEAM_NAME'] == team_b_nickname, 'PLAYOFF_RANK'].values[0]  
        # Convert to string
        team_a_playoff_rank = str(team_a_playoff_rank)
        team_b_playoff_rank = str(team_b_playoff_rank)

        # Get the win percentages of the teams
        team_a_win_percentage = league_standings_df.loc[league_standings_df['TEAM_NAME'] == team_a_nickname, 'WIN_PERCENTAGE'].values[0]
        team_b_win_percentage = league_standings_df.loc[league_standings_df['TEAM_NAME'] == team_b_nickname, 'WIN_PERCENTAGE'].values[0]  
        # Format at percentage and convert to string
        team_a_win_percentage = str((team_a_win_percentage * 100).round(2)) + '%'
        team_b_win_percentage = str((team_b_win_percentage * 100).round(2)) + '%'

        prediction_row = [index + 1,
                          team_a_nickname + ' | ' + team_b_nickname, 
                          team_a_league_record + ' | ' + team_b_league_record,
                          team_a_home_record + ' | ' + team_b_home_record,
                          team_a_road_record + ' | ' + team_b_road_record,
                          team_a_current_streak + ' | ' + team_b_current_streak,
                          team_a_playoff_rank + ' | ' + team_b_playoff_rank,
                          team_a_win_percentage + ' | ' + team_b_win_percentage,
                          predicted_winner, 
                          probability]

        # Add the prediction row to the predictions table
        predictions.add_row(prediction_row)

        # Add the prediction row to the predictions list
        predictions_list.append(prediction_row)

    print('Done!\n')

    print(predictions)

    # Put the predictions list into a dataframe
    predictions_list_df = pd.DataFrame(predictions_list)

    # Append the prediction rows to the testing Excel file
    append_to_excel(predictions_list_df, game_date)
    
# Function to predict the outcome of a game
def predict_game_outcome(team_a, team_b, combined_stats_df):  
  
    # Use the loaded model to make the prediction 
    prediction = model.predict(combined_stats_df)

    # Get the probability of the prediction
    prediction_probability = model.predict_proba(combined_stats_df)  

    if prediction == 1:

        probability = (prediction_probability[:, 1] * 100).round(2)
    
        prediction_result = {'Winner':[team_a], 'Probability':[str(probability)[1:-1] + '%']}

    else:

        probability = (prediction_probability[:, 0] * 100).round(2)

        prediction_result = {'Winner':[team_b], 'Probability':[str(probability)[1:-1] + '%']}
    
    return prediction_result 

# Function to append the prediction rows to the testing Excel file
def append_to_excel(predictions_list_df, game_date):

    # The column names to be used
    column_names = ['#', 'Home Team | Visitor Team', 'League Record', 
                    'Home Record', 'Road Record', 'Current Streak', 'Playoff Rank', 
                    'Win Percentage', 'Predicted Winner', 'Probability']

    # Assign the column names to the data frame
    predictions_list_df.columns = column_names

    # Add a column with the game date
    predictions_list_df['Game Date'] = game_date

    # Rearrange the columns to start with the game date
    predictions_list_df = predictions_list_df[['Game Date', '#', 'Home Team | Visitor Team', 'League Record', 
                                               'Home Record', 'Road Record', 'Current Streak', 'Playoff Rank', 
                                               'Win Percentage', 'Predicted Winner', 'Probability']]

    # Drop the index or number column
    predictions_list_df = predictions_list_df.drop(columns = ['#'])

    # Load the Excel workbook
    work_book = openpyxl.load_workbook('tests/tests.xlsx')

    # Create an Excel writer and assing the workbook to it
    excel_writer = pd.ExcelWriter('tests/tests.xlsx', engine = 'openpyxl', mode = 'a')
    excel_writer.book = work_book

    # Get the worksheets
    excel_writer.sheets = dict((work_sheet.title, work_sheet) for work_sheet in work_book.worksheets)

    # The row in the Excel worksheet where appending the prediction rows will start
    start_row = 0 

    # Loop through the rows and find where appending the prediction rows should start
    for row in work_book['Predictions']:

        # If the cell value is empty
        if work_book['Predictions'].cell(row[0].row, 1).value == None and work_book['Predictions'].cell(row[0].row, 2).value == None:

            # Get the row index and exit the loop
            start_row = row[0].row - 1
            break

    # Loop through the prediction rows and check if the record already exists in the worksheet
    # to avoid making duplicate entries
    for index, row in predictions_list_df.iterrows():

        # Loop through the worksheet rows
        for row in work_book['Predictions']:

            # If the record exists
            if predictions_list_df['Game Date'][index] == work_book['Predictions'].cell(row[0].row, 1).value and predictions_list_df['Home Team | Visitor Team'][index] == work_book['Predictions'].cell(row[0].row, 2).value:

                # Remove the record from the data frame and exit the inner loop
                predictions_list_df = predictions_list_df.drop(index)
                break

    # Check the number of prediction rows left and proceed to appending to the testing Excel file
    # If there are rows left to append                  
    if len(predictions_list_df) > 0:

        # Append the prediction rows to the worksheet
        predictions_list_df.to_excel(excel_writer, 'Predictions', startrow = start_row, index = False, header = False)
    
        # Save and close the Excel writer
        excel_writer.save()
        excel_writer.close()

        # If the number of prediction rows appended is 1
        if len(predictions_list_df) == 1:
     
            print('\n1 prediction row added to testing Excel file and saved!')

        # If the number of prediction rows appended is more than one
        else:

            print('\n' + str(len(predictions_list_df)) + ' prediction rows added to testing Excel file and saved!')    

    # If there are no rows left to append
    else:

        print('\nNo prediction rows were added to testing Excel file!')     

if __name__ == '__main__':

    # Get the games schedule and start the prediction process
    # The date should be in the format 'mm/dd/yy'
    get_games_schedule('5/19/2021', '2020-21')        